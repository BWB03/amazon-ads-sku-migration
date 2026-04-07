#!/usr/bin/env python3
"""
Amazon Advertising - Add SKU to Existing Ad Groups

Finds all ad groups/campaigns containing specified SKUs, then generates
bulk upload rows to add a new SKU to those same ad groups.

Use case: You have a SKU that needs to be present in certain campaigns
(e.g., an FBA SKU tied to a running deal) but can't delete it. This script
adds it to all the ad groups where related SKUs already exist.

Usage:
    python add_sku_to_adgroups.py <bulk_file.xlsx> [options]

Requirements:
    pip install openpyxl
"""

import os
import sys
import argparse
from openpyxl import Workbook, load_workbook


# ── Sheet Configurations ───────────────────────────────────────────────────

SHEET_CONFIGS = {
    "Sponsored Products Campaigns": {
        "product_value": "Sponsored Products",
        "entity_col": 2,       # B - Entity
        "operation_col": 3,    # C - Operation
        "campaign_id_col": 4,  # D - Campaign ID
        "ad_group_id_col": 5,  # E - Ad Group ID
        "state_col": 18,       # R - State
        "sku_col": 22,         # V - SKU
    },
    "Sponsored Display Campaigns": {
        "product_value": "Sponsored Display",
        "entity_col": 2,       # B - Entity
        "operation_col": 3,    # C - Operation
        "campaign_id_col": 4,  # D - Campaign ID
        "ad_group_id_col": 6,  # F - Ad Group ID (NOT column E like SP!)
        "state_col": 16,       # P - State
        "sku_col": 22,         # V - SKU
    },
}


def read_product_ads(ws, config):
    """Read all Product Ad rows from a worksheet."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        entity = row[config["entity_col"] - 1].value
        if entity != "Product Ad":
            continue

        sku = row[config["sku_col"] - 1].value
        campaign_id = row[config["campaign_id_col"] - 1].value
        ad_group_id = row[config["ad_group_id_col"] - 1].value
        state = row[config["state_col"] - 1].value

        if sku is not None and str(sku).strip() == "":
            sku = None

        rows.append({
            "campaign_id": str(campaign_id).strip() if campaign_id else None,
            "ad_group_id": str(ad_group_id).strip() if ad_group_id else None,
            "sku": str(sku).strip() if sku else None,
            "state": state,
        })
    return rows


def identify_adgroups_to_add(product_ads, search_skus, new_sku):
    """
    Find all ad groups containing any of the search SKUs, then generate
    new rows to add new_sku to those ad groups.

    Returns:
        new_rows: list of dicts with campaign_id, ad_group_id, new_sku
        matched_ad_groups: set of (campaign_id, ad_group_id) that matched
        skipped_duplicates: list of ad groups where new_sku already exists
    """
    # Normalize search SKUs for case-insensitive matching
    search_skus_upper = {s.upper() for s in search_skus}

    # Build set of existing (ad_group_id, sku) pairs to detect duplicates
    existing_pairs = set()
    for row in product_ads:
        if row["sku"] and row["ad_group_id"]:
            existing_pairs.add((row["ad_group_id"], row["sku"]))

    # Find all unique (campaign_id, ad_group_id) pairs that contain a search SKU
    matched_ad_groups = {}
    for row in product_ads:
        if row["sku"] and row["sku"].upper() in search_skus_upper:
            key = (row["campaign_id"], row["ad_group_id"])
            if key not in matched_ad_groups:
                matched_ad_groups[key] = row["sku"]  # track which SKU matched

    new_rows = []
    skipped_duplicates = []

    for (campaign_id, ad_group_id), matched_sku in matched_ad_groups.items():
        # Check if new_sku already exists in this ad group
        if (ad_group_id, new_sku) in existing_pairs:
            skipped_duplicates.append(
                f"Ad group {ad_group_id} (campaign {campaign_id}) - already has {new_sku}"
            )
            continue

        new_rows.append({
            "campaign_id": campaign_id,
            "ad_group_id": ad_group_id,
            "matched_sku": matched_sku,
            "new_sku": new_sku,
        })

    return new_rows, set(matched_ad_groups.keys()), skipped_duplicates


def write_output_file(filepath, sheet_data, source_headers):
    """Write an output xlsx with new Product Ad rows."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, (new_rows, config) in sheet_data.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = source_headers[sheet_name]

        for col_idx, header_val in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header_val)

        if not new_rows:
            continue

        for i, row_data in enumerate(new_rows, start=2):
            ws.cell(row=i, column=1, value=config["product_value"])
            ws.cell(row=i, column=config["entity_col"], value="Product Ad")
            ws.cell(row=i, column=config["operation_col"], value="Create")
            ws.cell(row=i, column=config["campaign_id_col"], value=row_data["campaign_id"])
            ws.cell(row=i, column=config["ad_group_id_col"], value=row_data["ad_group_id"])
            ws.cell(row=i, column=config["state_col"], value="enabled")
            ws.cell(row=i, column=config["sku_col"], value=row_data["new_sku"])

    wb.save(filepath)
    print(f"  Saved: {filepath}")


def main():
    parser = argparse.ArgumentParser(
        description="Add a SKU to all ad groups where specified SKUs exist.",
        epilog="Example: python add_sku_to_adgroups.py bulk_download.xlsx",
    )
    parser.add_argument(
        "source_file",
        help="Path to your Amazon Advertising bulk download .xlsx file",
    )
    parser.add_argument(
        "--search-skus",
        nargs="+",
        required=True,
        help="SKUs to search for (space-separated)",
    )
    parser.add_argument(
        "--add-sku",
        required=True,
        help="SKU to add to matched ad groups",
    )
    parser.add_argument(
        "--test-campaign",
        default=None,
        help="Campaign ID to use for the test file (optional)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory (defaults to same directory as source file)",
    )
    parser.add_argument(
        "--full-only",
        action="store_true",
        help="Only generate the full file, skip the test file",
    )

    args = parser.parse_args()

    if not os.path.exists(args.source_file):
        print(f"Error: File not found: {args.source_file}")
        sys.exit(1)

    output_dir = args.output_dir or os.path.dirname(os.path.abspath(args.source_file))
    full_output = os.path.join(output_dir, "Add_SKU_FULL.xlsx")
    test_output = os.path.join(output_dir, "Add_SKU_TEST.xlsx")

    print("=" * 70)
    print("Amazon Advertising - Add SKU to Ad Groups")
    print("=" * 70)
    print(f"\n  Source file:  {args.source_file}")
    print(f"  Search SKUs:  {args.search_skus}")
    print(f"  SKU to add:   {args.add_sku}")
    print(f"  Output dir:   {output_dir}")
    print("\nLoading workbook...")

    wb = load_workbook(args.source_file, data_only=True)

    all_sheet_data_full = {}
    source_headers = {}
    total_new = 0
    all_new_rows_with_sheet = []

    for sheet_name, config in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"\n  Skipping '{sheet_name}' (not found in workbook)")
            continue

        print(f"\n{'─' * 50}")
        print(f"Processing: {sheet_name}")
        print(f"{'─' * 50}")

        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        source_headers[sheet_name] = headers

        product_ads = read_product_ads(ws, config)
        print(f"  Found {len(product_ads)} existing Product Ad rows")

        new_rows, matched, skipped_dupes = identify_adgroups_to_add(
            product_ads, args.search_skus, args.add_sku
        )

        print(f"  Ad groups matched (contain search SKUs): {len(matched)}")
        print(f"  New rows to create: {len(new_rows)}")

        if skipped_dupes:
            print(f"  Skipped (SKU already in ad group): {len(skipped_dupes)}")
            for s in skipped_dupes:
                print(f"    {s}")

        if new_rows:
            print(f"\n  Ad groups receiving '{args.add_sku}':")
            for r in new_rows:
                print(f"    Campaign {r['campaign_id']} / Ad Group {r['ad_group_id']}  (matched via {r['matched_sku']})")

        all_sheet_data_full[sheet_name] = (new_rows, config)
        for r in new_rows:
            all_new_rows_with_sheet.append((sheet_name, r))
        total_new += len(new_rows)

    if total_new == 0:
        print("\nNo new rows to create. The SKU may already exist in all matched ad groups.")
        wb.close()
        sys.exit(0)

    # Determine test campaign
    test_campaign_id = args.test_campaign
    if not test_campaign_id and not args.full_only:
        campaign_counts = {}
        for sheet_name, r in all_new_rows_with_sheet:
            cid = r["campaign_id"]
            if cid:
                campaign_counts[cid] = campaign_counts.get(cid, 0) + 1
        if campaign_counts:
            test_campaign_id = min(campaign_counts, key=campaign_counts.get)
            print(f"\n  Auto-selected test campaign: {test_campaign_id} "
                  f"({campaign_counts[test_campaign_id]} row(s))")

    # Build test data
    all_sheet_data_test = {}
    total_test = 0
    if not args.full_only and test_campaign_id:
        for sheet_name, (new_rows, config) in all_sheet_data_full.items():
            test_rows = [r for r in new_rows if r["campaign_id"] == test_campaign_id]
            all_sheet_data_test[sheet_name] = (test_rows, config)
            total_test += len(test_rows)

    # Write output files
    print(f"\n{'=' * 70}")
    print("Writing output files...")
    print(f"{'=' * 70}")

    write_output_file(full_output, all_sheet_data_full, source_headers)
    print(f"    {total_new} new Product Ad rows")

    if not args.full_only and total_test > 0:
        write_output_file(test_output, all_sheet_data_test, source_headers)
        print(f"    {total_test} new Product Ad rows (campaign: {test_campaign_id})")

    wb.close()

    # Summary
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total new Product Ad rows (FULL): {total_new}")
    if not args.full_only and total_test > 0:
        print(f"  Total new Product Ad rows (TEST): {total_test}")
    print(f"\n  Next steps:")
    if not args.full_only and total_test > 0:
        print(f"    1. Open the TEST file in Excel and review the rows")
        print(f"    2. Upload TEST file to Amazon Advertising bulk operations")
        print(f"    3. Verify the new Product Ad appears correctly")
        print(f"    4. Upload FULL file after test succeeds")
    else:
        print(f"    1. Open the FULL file in Excel and review the rows")
        print(f"    2. Upload to Amazon Advertising bulk operations")
        print(f"    3. Monitor the bulk upload processing report for errors")
    print(f"\nDone!")


if __name__ == "__main__":
    main()
