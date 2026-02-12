#!/usr/bin/env python3
"""
Amazon Advertising - Bulk SKU Migration Script

When transitioning from Amazon sticker barcodes to manufacturer barcodes,
every existing SKU needs a new Product Ad created with the updated SKU.

Per Amazon: "You can't update a SKU or ASIN, but you can pause or archive
the associated product ad entity and then create a new one with the SKU
or ASIN you want to add."

This script reads your existing Amazon Advertising bulk download file,
identifies all Product Ad rows, and generates upload-ready files with
new Product Ad rows using your updated SKUs.

Usage:
    python generate_sku_migration.py <bulk_file.xlsx> [options]

Requirements:
    pip install openpyxl
"""

import os
import sys
import argparse
from openpyxl import Workbook, load_workbook


# ── Sheet Configurations ───────────────────────────────────────────────────
# Maps sheet names to their column layouts.
# Amazon's bulk sheets have different column positions for SP vs SD.

SHEET_CONFIGS = {
    "Sponsored Products Campaigns": {
        "product_value": "Sponsored Products",
        "entity_col": 2,       # B - Entity
        "operation_col": 3,    # C - Operation
        "campaign_id_col": 4,  # D - Campaign ID
        "ad_group_id_col": 5,  # E - Ad Group ID
        "state_col": 18,       # R - State
        "sku_col": 22,         # V - SKU
    },
    "Sponsored Display Campaigns": {
        "product_value": "Sponsored Display",
        "entity_col": 2,       # B - Entity
        "operation_col": 3,    # C - Operation
        "campaign_id_col": 4,  # D - Campaign ID
        "ad_group_id_col": 6,  # F - Ad Group ID (NOT column E like SP!)
        "state_col": 16,       # P - State
        "sku_col": 22,         # V - SKU
    },
}


def build_sku_transform(mode, suffix):
    """
    Return a function that transforms an old SKU into a new SKU.

    Modes:
        'suffix' - Appends a suffix (e.g., 'A') to the existing SKU
        'custom' - Reserved for future use (custom mapping from CSV, etc.)
    """
    if mode == "suffix":
        def transform(sku):
            return str(sku) + suffix
        return transform
    else:
        raise ValueError(f"Unknown transform mode: {mode}")


def should_skip_sku(sku, new_sku_suffix):
    """Return True if this SKU should be excluded from migration."""
    if not sku:
        return True, "empty"
    sku_str = str(sku).strip()
    if not sku_str:
        return True, "empty"
    # Skip SKUs that already end with the suffix (avoid double-appending)
    if sku_str.upper().endswith(new_sku_suffix.upper()):
        return True, "already_has_suffix"
    # Skip FBA/missing error entries
    if "FBA" in sku_str.upper() or "MISSING" in sku_str.upper():
        return True, "fba_or_missing"
    return False, None


def read_product_ads(ws, config):
    """Read all Product Ad rows from a worksheet."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        entity = row[config["entity_col"] - 1].value
        if entity != "Product Ad":
            continue

        sku = row[config["sku_col"] - 1].value
        campaign_id = row[config["campaign_id_col"] - 1].value
        ad_group_id = row[config["ad_group_id_col"] - 1].value
        state = row[config["state_col"] - 1].value

        # Normalize empty strings to None
        if sku is not None and str(sku).strip() == "":
            sku = None

        rows.append({
            "campaign_id": str(campaign_id).strip() if campaign_id else None,
            "ad_group_id": str(ad_group_id).strip() if ad_group_id else None,
            "sku": str(sku).strip() if sku else None,
            "state": state,
        })
    return rows


def identify_new_rows(product_ads, transform_fn, suffix):
    """
    Determine which Product Ad rows need new variants.

    Returns:
        new_rows: list of dicts with campaign_id, ad_group_id, old_sku, new_sku
        skipped: dict of skip reasons with lists of SKUs
    """
    # Build set of existing (ad_group_id, sku) pairs to avoid duplicates
    existing_pairs = set()
    for row in product_ads:
        if row["sku"] and row["ad_group_id"]:
            existing_pairs.add((row["ad_group_id"], row["sku"]))

    new_rows = []
    skipped = {"already_has_suffix": [], "fba_or_missing": [], "duplicate": [], "empty": []}

    for row in product_ads:
        sku = row["sku"]

        skip, reason = should_skip_sku(sku, suffix)
        if skip:
            if reason:
                skipped[reason].append(sku or "(empty)")
            continue

        new_sku = transform_fn(sku)

        # Check if the new SKU already exists in this ad group
        if (row["ad_group_id"], new_sku) in existing_pairs:
            skipped["duplicate"].append(
                f"{sku} -> {new_sku} (ad group {row['ad_group_id']})"
            )
            continue

        new_rows.append({
            "campaign_id": row["campaign_id"],
            "ad_group_id": row["ad_group_id"],
            "old_sku": sku,
            "new_sku": new_sku,
        })

    return new_rows, skipped


def write_output_file(filepath, sheet_data, source_headers):
    """Write an output xlsx with new Product Ad rows."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, (new_rows, config) in sheet_data.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = source_headers[sheet_name]

        # Write header row (copied exactly from source file)
        for col_idx, header_val in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header_val)

        if not new_rows:
            continue

        # Write new Product Ad rows
        for i, row_data in enumerate(new_rows, start=2):
            ws.cell(row=i, column=1, value=config["product_value"])
            ws.cell(row=i, column=config["entity_col"], value="Product Ad")
            ws.cell(row=i, column=config["operation_col"], value="Create")
            ws.cell(row=i, column=config["campaign_id_col"], value=row_data["campaign_id"])
            ws.cell(row=i, column=config["ad_group_id_col"], value=row_data["ad_group_id"])
            ws.cell(row=i, column=config["state_col"], value="enabled")
            ws.cell(row=i, column=config["sku_col"], value=row_data["new_sku"])

    wb.save(filepath)
    print(f"  Saved: {filepath}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Amazon Advertising bulk upload files for SKU migration.",
        epilog="Example: python generate_sku_migration.py bulk_download.xlsx --suffix A",
    )
    parser.add_argument(
        "source_file",
        help="Path to your Amazon Advertising bulk download .xlsx file",
    )
    parser.add_argument(
        "--suffix",
        default="A",
        help="Suffix to append to each SKU (default: 'A')",
    )
    parser.add_argument(
        "--test-campaign",
        default=None,
        help="Campaign ID to use for the test file (optional). "
             "If not provided, the script will pick the campaign with the fewest new rows.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory (defaults to same directory as source file)",
    )
    parser.add_argument(
        "--full-only",
        action="store_true",
        help="Only generate the full file, skip the test file",
    )

    args = parser.parse_args()

    if not os.path.exists(args.source_file):
        print(f"Error: File not found: {args.source_file}")
        sys.exit(1)

    output_dir = args.output_dir or os.path.dirname(os.path.abspath(args.source_file))
    full_output = os.path.join(output_dir, "SKU_Migration_FULL.xlsx")
    test_output = os.path.join(output_dir, "SKU_Migration_TEST.xlsx")

    transform_fn = build_sku_transform("suffix", args.suffix)

    print("=" * 70)
    print("Amazon Advertising - SKU Migration Script")
    print("=" * 70)
    print(f"\n  Source file: {args.source_file}")
    print(f"  SKU suffix:  '{args.suffix}'")
    print(f"  Output dir:  {output_dir}")
    print("\nLoading workbook...")

    wb = load_workbook(args.source_file, data_only=True)

    all_sheet_data_full = {}
    all_sheet_data_test = {}
    source_headers = {}
    total_new = 0
    total_test = 0
    all_new_rows_with_sheet = []  # for auto-selecting test campaign

    for sheet_name, config in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"\n  Skipping '{sheet_name}' (not found in workbook)")
            continue

        print(f"\n{'─' * 50}")
        print(f"Processing: {sheet_name}")
        print(f"{'─' * 50}")

        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        source_headers[sheet_name] = headers

        product_ads = read_product_ads(ws, config)
        print(f"  Found {len(product_ads)} existing Product Ad rows")

        new_rows, skipped = identify_new_rows(product_ads, transform_fn, args.suffix)
        print(f"  New rows to create: {len(new_rows)}")

        if skipped["already_has_suffix"]:
            print(f"  Skipped (already ends in '{args.suffix}'): {len(skipped['already_has_suffix'])} rows")
        if skipped["fba_or_missing"]:
            print(f"  Skipped (FBA/missing): {len(skipped['fba_or_missing'])} rows")
        if skipped["duplicate"]:
            print(f"  Skipped (new SKU already exists): {len(skipped['duplicate'])} rows")

        # Print unique SKU mappings
        seen = set()
        print(f"\n  SKU Mappings:")
        for r in new_rows:
            key = (r["old_sku"], r["new_sku"])
            if key not in seen:
                seen.add(key)
                print(f"    {r['old_sku']:30s} -> {r['new_sku']}")

        all_sheet_data_full[sheet_name] = (new_rows, config)
        for r in new_rows:
            all_new_rows_with_sheet.append((sheet_name, r))
        total_new += len(new_rows)

    if total_new == 0:
        print("\nNo new rows to create. Nothing to do!")
        wb.close()
        sys.exit(0)

    # Determine test campaign
    test_campaign_id = args.test_campaign
    if not test_campaign_id and not args.full_only:
        # Auto-select: pick campaign with fewest new rows (simplest test)
        campaign_counts = {}
        for sheet_name, r in all_new_rows_with_sheet:
            cid = r["campaign_id"]
            if cid:
                campaign_counts[cid] = campaign_counts.get(cid, 0) + 1
        if campaign_counts:
            test_campaign_id = min(campaign_counts, key=campaign_counts.get)
            print(f"\n  Auto-selected test campaign: {test_campaign_id} "
                  f"({campaign_counts[test_campaign_id]} row(s))")

    # Build test data
    if not args.full_only and test_campaign_id:
        for sheet_name, (new_rows, config) in all_sheet_data_full.items():
            test_rows = [r for r in new_rows if r["campaign_id"] == test_campaign_id]
            all_sheet_data_test[sheet_name] = (test_rows, config)
            total_test += len(test_rows)

    # Write output files
    print(f"\n{'=' * 70}")
    print("Writing output files...")
    print(f"{'=' * 70}")

    write_output_file(full_output, all_sheet_data_full, source_headers)
    print(f"    {total_new} new Product Ad rows")

    if not args.full_only and total_test > 0:
        write_output_file(test_output, all_sheet_data_test, source_headers)
        print(f"    {total_test} new Product Ad rows (campaign: {test_campaign_id})")

    wb.close()

    # Summary
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total new Product Ad rows (FULL): {total_new}")
    if not args.full_only and total_test > 0:
        print(f"  Total new Product Ad rows (TEST): {total_test}")
    print(f"\n  Next steps:")
    if not args.full_only and total_test > 0:
        print(f"    1. Open the TEST file in Excel and review the rows")
        print(f"    2. Upload TEST file to Amazon Advertising bulk operations")
        print(f"    3. Verify the new Product Ad appears correctly (24-48 hrs)")
        print(f"    4. Upload FULL file after test succeeds")
    else:
        print(f"    1. Open the FULL file in Excel and review the rows")
        print(f"    2. Upload to Amazon Advertising bulk operations")
        print(f"    3. Monitor the bulk upload processing report for errors")
    print(f"\n  NOTE: Sponsored Brands campaigns use ASINs (not SKUs) and")
    print(f"  typically don't need updates when changing barcode types.")
    print(f"\nDone!")


if __name__ == "__main__":
    main()
