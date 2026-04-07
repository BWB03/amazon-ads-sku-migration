#!/usr/bin/env python3
"""
Amazon Advertising - Negative Keyword Generator

Reads a curated list of search terms (with match type recommendations) and
an Amazon Ads bulk download file, then outputs upload-ready bulk files with
negative keyword rows added to the correct campaigns and sheets.

Supports Sponsored Products and Sponsored Brands (SBV) campaigns.

Usage:
    python generate_negative_keywords.py <bulk_file.xlsx> --negates <negate_list.csv>

Requirements:
    pip install openpyxl
"""

import os
import sys
import csv
import argparse
from openpyxl import Workbook, load_workbook

from bulk_utils import (
    NEGATIVE_KEYWORD_SHEETS,
    SHEET_PRODUCT_VALUES,
    find_columns_by_header,
    find_column,
    build_campaign_lookup,
    read_existing_negatives,
    auto_select_test_campaign,
)


def parse_negate_csv(filepath):
    """
    Read the negate list CSV or XLSX file.

    Expected columns: search_term, campaign_name, match_type_recommendation, scope
    Additional columns (spend, orders, acos_percent, reason) are ignored.

    Returns list of dicts.
    """
    negates = []

    if filepath.lower().endswith(".xlsx"):
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[header] = row[i]
            if row_dict.get("search_term") and row_dict.get("campaign_name"):
                negates.append(row_dict)
        wb.close()
    else:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("search_term") and row.get("campaign_name"):
                    negates.append(row)

    return negates


def match_campaign(campaign_name, campaign_lookup):
    """
    Match a campaign name from the negate list to the bulk file lookup.

    Tries exact match, then case-insensitive match.
    Returns (matched_name, campaign_data) or (None, None).
    """
    # Exact match
    if campaign_name in campaign_lookup:
        return campaign_name, campaign_lookup[campaign_name]

    # Case-insensitive match
    name_lower = campaign_name.lower().strip()
    for lookup_name, data in campaign_lookup.items():
        if lookup_name.lower().strip() == name_lower:
            return lookup_name, data

    return None, None


def build_negative_rows(negates, campaign_lookup, existing_campaign_negs, existing_adgroup_negs):
    """
    Build negative keyword rows from the negate list.

    Returns:
        new_rows: list of row dicts ready for output
        warnings: list of warning strings
        stats: dict of counts
    """
    new_rows = []
    warnings = []
    stats = {"matched": 0, "unmatched": 0, "duplicates": 0, "created": 0}

    # Track what we're adding to avoid duplicates within the negate list itself
    adding_campaign = set()  # (campaign_id, keyword_lower, match_type)
    adding_adgroup = set()   # (ad_group_id, keyword_lower, match_type)

    for negate in negates:
        search_term = str(negate["search_term"]).strip()
        csv_campaign_name = str(negate["campaign_name"]).strip()
        match_type = str(negate.get("match_type_recommendation", "negativePhrase")).strip()
        scope = str(negate.get("scope", "campaign")).strip().lower()

        matched_name, campaign_data = match_campaign(csv_campaign_name, campaign_lookup)

        if not campaign_data:
            stats["unmatched"] += 1
            warnings.append(f"Campaign not found in bulk file: '{csv_campaign_name}'")
            continue

        stats["matched"] += 1
        campaign_id = campaign_data["campaign_id"]
        sheet_name = campaign_data["sheet_name"]

        if scope == "adgroup":
            entity = "Negative keyword"
            ad_groups = campaign_data.get("ad_groups", {})
            if not ad_groups:
                warnings.append(
                    f"No ad groups found for '{csv_campaign_name}' — using campaign-level instead"
                )
                entity = "Campaign negative keyword"
                scope = "campaign"

        if scope == "campaign":
            entity = "Campaign negative keyword"

            dedup_key = (campaign_id, search_term.lower(), match_type)
            if dedup_key in existing_campaign_negs or dedup_key in adding_campaign:
                stats["duplicates"] += 1
                continue

            adding_campaign.add(dedup_key)
            new_rows.append({
                "campaign_id": campaign_id,
                "campaign_name": matched_name,
                "ad_group_id": None,
                "ad_group_name": None,
                "entity": entity,
                "keyword_text": search_term,
                "match_type": match_type,
                "sheet_name": sheet_name,
            })
            stats["created"] += 1

        elif scope == "adgroup":
            for ag_name, ag_id in ad_groups.items():
                dedup_key = (ag_id, search_term.lower(), match_type)
                if dedup_key in existing_adgroup_negs or dedup_key in adding_adgroup:
                    stats["duplicates"] += 1
                    continue

                adding_adgroup.add(dedup_key)
                new_rows.append({
                    "campaign_id": campaign_id,
                    "campaign_name": matched_name,
                    "ad_group_id": ag_id,
                    "ad_group_name": ag_name,
                    "entity": entity,
                    "keyword_text": search_term,
                    "match_type": match_type,
                    "sheet_name": sheet_name,
                })
                stats["created"] += 1

    return new_rows, warnings, stats


def write_output_file(filepath, rows_by_sheet, source_headers, header_maps):
    """Write an output xlsx with negative keyword rows."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in rows_by_sheet.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = source_headers[sheet_name]

        # Write header row (copied exactly from source file)
        for col_idx, header_val in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header_val)

        if not rows:
            continue

        hmap = header_maps[sheet_name]
        product_value = SHEET_PRODUCT_VALUES.get(sheet_name, "Sponsored Products")

        for i, row_data in enumerate(rows, start=2):
            # Product (column A, or by header)
            product_col = find_column(hmap, "Product") or 1
            ws.cell(row=i, column=product_col, value=product_value)

            entity_col = find_column(hmap, "Entity")
            if entity_col:
                ws.cell(row=i, column=entity_col, value=row_data["entity"])

            op_col = find_column(hmap, "Operation")
            if op_col:
                ws.cell(row=i, column=op_col, value="Create")

            cid_col = find_column(hmap, "Campaign ID")
            if cid_col:
                ws.cell(row=i, column=cid_col, value=row_data["campaign_id"])

            cname_col = find_column(
                hmap, "Campaign Name", "Campaign Name (Informational only)"
            )
            if cname_col:
                ws.cell(row=i, column=cname_col, value=row_data["campaign_name"])

            # Ad Group columns only for ad-group-level negatives
            if row_data.get("ad_group_id"):
                agid_col = find_column(hmap, "Ad Group ID")
                if agid_col:
                    ws.cell(row=i, column=agid_col, value=row_data["ad_group_id"])
                agname_col = find_column(
                    hmap, "Ad Group Name", "Ad Group Name (Informational only)"
                )
                if agname_col:
                    ws.cell(row=i, column=agname_col, value=row_data["ad_group_name"])

            kw_col = find_column(hmap, "Keyword Text", "Keyword or Product Targeting")
            if kw_col:
                ws.cell(row=i, column=kw_col, value=row_data["keyword_text"])

            mt_col = find_column(hmap, "Match Type")
            if mt_col:
                ws.cell(row=i, column=mt_col, value=row_data["match_type"])

            state_col = find_column(hmap, "State")
            if state_col:
                ws.cell(row=i, column=state_col, value="enabled")

    wb.save(filepath)
    print(f"  Saved: {filepath}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Amazon Advertising bulk upload files for negative keywords.",
        epilog="Example: python generate_negative_keywords.py bulk_download.xlsx --negates negate_list.csv",
    )
    parser.add_argument(
        "source_file",
        help="Path to your Amazon Advertising bulk download .xlsx file",
    )
    parser.add_argument(
        "--negates",
        required=True,
        help="Path to negate list CSV or XLSX "
        "(columns: search_term, campaign_name, match_type_recommendation, scope)",
    )
    parser.add_argument(
        "--test-campaign",
        default=None,
        help="Campaign ID for the test file (optional, auto-selects if not provided)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory (defaults to same directory as source file)",
    )
    parser.add_argument(
        "--full-only",
        action="store_true",
        help="Only generate the full file, skip the test file",
    )

    args = parser.parse_args()

    if not os.path.exists(args.source_file):
        print(f"Error: File not found: {args.source_file}")
        sys.exit(1)

    if not os.path.exists(args.negates):
        print(f"Error: Negate list not found: {args.negates}")
        sys.exit(1)

    output_dir = args.output_dir or os.path.dirname(os.path.abspath(args.source_file))
    full_output = os.path.join(output_dir, "Negative_Keywords_FULL.xlsx")
    test_output = os.path.join(output_dir, "Negative_Keywords_TEST.xlsx")

    print("=" * 70)
    print("Amazon Advertising - Negative Keyword Generator")
    print("=" * 70)
    print(f"\n  Source file:  {args.source_file}")
    print(f"  Negate list:  {args.negates}")
    print(f"  Output dir:   {output_dir}")

    # Load negate list
    print("\nLoading negate list...")
    negates = parse_negate_csv(args.negates)
    print(f"  {len(negates)} search terms to negate")

    # Load bulk workbook
    print("\nLoading bulk workbook...")
    wb = load_workbook(args.source_file, data_only=True)

    # Build campaign lookup from bulk file
    print("\nBuilding campaign lookup...")
    campaign_lookup = build_campaign_lookup(wb)
    print(f"  Found {len(campaign_lookup)} campaigns across bulk file")

    # Read existing negatives for dedup
    existing_campaign_negs, existing_adgroup_negs = read_existing_negatives(wb)
    print(f"  Existing campaign-level negatives: {len(existing_campaign_negs)}")
    print(f"  Existing ad-group-level negatives: {len(existing_adgroup_negs)}")

    # Collect source headers and header maps for output writing
    source_headers = {}
    header_maps = {}
    for sheet_name in NEGATIVE_KEYWORD_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            source_headers[sheet_name] = [
                cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            header_maps[sheet_name] = find_columns_by_header(ws)

    # Build negative keyword rows
    print(f"\n{'─' * 50}")
    print("Processing negatives...")
    print(f"{'─' * 50}")

    new_rows, warnings, stats = build_negative_rows(
        negates, campaign_lookup, existing_campaign_negs, existing_adgroup_negs
    )

    # Print warnings (deduplicated)
    if warnings:
        unique_warnings = list(dict.fromkeys(warnings))
        print(f"\n  Warnings ({len(unique_warnings)}):")
        for w in unique_warnings:
            print(f"    WARNING: {w}")

    print(f"\n  Campaigns matched:    {stats['matched']}")
    print(f"  Campaigns not found:  {stats['unmatched']}")
    print(f"  Duplicates skipped:   {stats['duplicates']}")
    print(f"  New rows to create:   {stats['created']}")

    if not new_rows:
        print("\nNo new negative keyword rows to create. Nothing to do!")
        wb.close()
        sys.exit(0)

    # Group rows by sheet
    rows_by_sheet_full = {}
    for sheet_name in source_headers:
        rows_by_sheet_full[sheet_name] = [
            r for r in new_rows if r["sheet_name"] == sheet_name
        ]

    # Print breakdown by sheet and match type
    for sheet_name, rows in rows_by_sheet_full.items():
        if rows:
            phrase_count = sum(1 for r in rows if r["match_type"] == "negativePhrase")
            exact_count = sum(1 for r in rows if r["match_type"] == "negativeExact")
            print(f"\n  {sheet_name}:")
            print(f"    negativePhrase: {phrase_count}")
            print(f"    negativeExact:  {exact_count}")

    # Print all negatives being added
    print(f"\n  Negative keywords to add:")
    for r in new_rows:
        scope_label = "campaign" if r["entity"] == "Campaign negative keyword" else "ad group"
        print(
            f"    [{r['match_type']:16s}] "
            f"{r['keyword_text']:40s} -> {r['campaign_name'][:50]} ({scope_label})"
        )

    # Test campaign selection
    test_campaign_id = args.test_campaign
    if not test_campaign_id and not args.full_only:
        test_campaign_id = auto_select_test_campaign(new_rows)
        if test_campaign_id:
            test_count = sum(1 for r in new_rows if r["campaign_id"] == test_campaign_id)
            test_camp_name = next(
                (r["campaign_name"] for r in new_rows if r["campaign_id"] == test_campaign_id),
                "",
            )
            print(f"\n  Auto-selected test campaign: {test_campaign_id}")
            print(f"    Name: {test_camp_name}")
            print(f"    Rows: {test_count}")

    # Build test data
    rows_by_sheet_test = {}
    total_test = 0
    if not args.full_only and test_campaign_id:
        for sheet_name in source_headers:
            test_rows = [
                r for r in new_rows
                if r["sheet_name"] == sheet_name and r["campaign_id"] == test_campaign_id
            ]
            rows_by_sheet_test[sheet_name] = test_rows
            total_test += len(test_rows)

    # Write output files
    print(f"\n{'=' * 70}")
    print("Writing output files...")
    print(f"{'=' * 70}")

    write_output_file(full_output, rows_by_sheet_full, source_headers, header_maps)
    print(f"    {len(new_rows)} negative keyword rows")

    if not args.full_only and total_test > 0:
        write_output_file(test_output, rows_by_sheet_test, source_headers, header_maps)
        print(f"    {total_test} negative keyword rows (campaign: {test_campaign_id})")

    wb.close()

    # Summary
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total negative keyword rows (FULL): {len(new_rows)}")
    if not args.full_only and total_test > 0:
        print(f"  Total negative keyword rows (TEST): {total_test}")
    print(f"\n  Next steps:")
    if not args.full_only and total_test > 0:
        print(f"    1. Open the TEST file in Excel and review the rows")
        print(f"    2. Upload TEST file to Amazon Advertising bulk operations")
        print(f"    3. Verify negatives appear correctly in Campaign Manager")
        print(f"    4. Upload FULL file after test succeeds")
    else:
        print(f"    1. Open the FULL file in Excel and review the rows")
        print(f"    2. Upload to Amazon Advertising bulk operations")
        print(f"    3. Monitor the bulk upload processing report for errors")
    print(f"\nDone!")


if __name__ == "__main__":
    main()
